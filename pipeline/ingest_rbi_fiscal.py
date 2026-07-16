"""RBI Handbook of Statistics on Indian States 2025 -> state fiscal vertical (5 metrics).

Source: Reserve Bank of India, Handbook of Statistics on Indian States 2025
(released Dec-2025). Five state-level metrics derived from five HBS tables:

  econ_percapita_nsdp_rbi   Per-capita NSDP, RBI HBS Table 19, latest FY, current prices (Rs)
                            -- ADDITIVE: distinct from the existing MoSPI
                               `econ_percapita_nsdp` metric; this is the RBI FY-latest series.
  gsdp_growth               YoY nominal GSDP growth, RBI HBS Table 21 two latest FYs (%)
  fiscal_deficit_pct_gsdp   Gross Fiscal Deficit (T164) / GSDP (T21) x 100 (%)
  own_tax_pct_gsdp          Own Tax Revenue (T168) / GSDP (T21) x 100 (%)
  outstanding_debt_pct_gsdp Outstanding Liabilities (T176, end-March stock) / GSDP (T21) x 100 (%)

UNIT ALIGNMENT (critical): T21 GSDP is in Rs LAKH; T164/T168/T176 are in Rs CRORE.
1 crore = 100 lakh, so GSDP_crore = T21_lakh / 100. All ratios are computed in a
consistent fiscal year: for each state we take the LATEST fiscal year for which BOTH
the numerator table AND the T21 GSDP have a usable value (the latest FY may carry the
tag Actual / RE / BE -- recorded in each metric's methodology and per-value estimated flag).

YEARS (RBI HBS 2025):
  T19 / T21 latest published FY = 2024-25 (some states only to 2023-24; '-' = Not Available).
  T164 / T168 columns end at 2024-25 (BE); 2023-24 is (RE); 2022-23 is (A)ctual.
  T176 outstanding-liabilities stock latest = as at end-March 2025 (= close of FY2024-25),
       with a clean "Outstanding Liabilities" total column (formula 20 = (14 to 19)).

State level only (RBI publishes state/UT government finances; no district fiscal series).
Aggregate rows ('All India', 'All States and UTs') and any name that does not resolve to a
region_keys state code are skipped and reported.

Run: pipeline/.venv/bin/python pipeline/ingest_rbi_fiscal.py
"""
import os, sqlite3
import openpyxl
from region_match import RegionMatcher, upsert_metric, log_load, DB

PIPE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(PIPE, "raw-new", "economy")

SOURCE = "Reserve Bank of India, Handbook of Statistics on Indian States 2025"
URL = "https://www.rbi.org.in/Scripts/AnnualPublications.aspx?head=Handbook+of+Statistics+on+Indian+States"
LICENSE = "RBI (Reserve Bank of India) publication; reproduction with acknowledgement"
FETCHED = "2026-07-01T00:00:00Z"      # scout fetch date for the HBS 2025 xlsx files

# Aggregate / non-region labels to always skip.
AGG = {"all india", "all states and uts", "all states and ut", "all-india", "india"}

# Source-name -> canonical form the RegionMatcher understands.
# RBI fiscal tables (T164/T168/T176) print Delhi as "NCT Delhi"; RegionMatcher only
# knows "nct of delhi". The "*" on "Jammu & Kashmir*" is stripped by norm() already.
STATE_ALIASES = {
    "nct delhi": "Delhi",
}

FILES = {
    "T19": "rbi_hbs2025_T19_percapita_nsdp_current.xlsx",
    "T21": "rbi_hbs2025_T21_gsdp_current.xlsx",
    "T164": "rbi_hbs2025_T164_gross_fiscal_deficit.xlsx",
    "T168": "rbi_hbs2025_T168_own_tax_revenue.xlsx",
    "T176": "rbi_hbs2025_T176_outstanding_liabilities.xlsx",
}

# Fiscal-year labelling for the latest few years of the budgetary tables (T164/T168).
# The numeric latest FY is 2024-25; estimate type varies by column suffix in the header.
ESTIMATE_TAG = {"2024-25": "BE", "2023-24": "RE", "2022-23": "A"}


def num(v):
    """Coerce a cell to a float, treating '-'/blank/'NA' as missing (None)."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "-", "–", "NA", "N.A.", "n.a.", "..."):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def fy_end_year(fy_label):
    """'2024-25' -> 2024 (store the STARTING calendar year of the fiscal year)."""
    return int(str(fy_label).split("-")[0])


def find_header(ws, max_scan=8):
    """Return (header_row, name_col) by locating the row whose text contains 'State'."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, min(ws.max_column, 6) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "state" in v.lower() and "union" in v.lower():
                return r, c
    # fallback: any cell containing 'State/Union'
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, min(ws.max_column, 6) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower().startswith("state"):
                return r, c
    raise RuntimeError(f"header row not found in sheet {ws.title!r}")


def year_columns(ws, header_row, name_col):
    """Map fiscal-year label -> column index for a T19/T21/T164/T168-style sheet.

    Year labels may sit on the header row (T164/T168) or on the row just below the
    'Base: 2011-12' banner (T19/T21). We scan the header row and the next row for
    cells matching a YYYY-YY fiscal-year pattern, then pick whichever row yields the
    MOST year columns -- this rejects the spurious single 'Base: 2011-12' banner cell
    on T19/T21 (which would otherwise masquerade as the only year column) in favour of
    the real per-year label row beneath it.
    """
    import re
    fy_re = re.compile(r"(\d{4})-(\d{2})")
    best_cols = {}
    for scan_row in (header_row, header_row + 1, header_row + 2):
        cols = {}
        for c in range(name_col + 1, ws.max_column + 1):
            v = ws.cell(row=scan_row, column=c).value
            if isinstance(v, str):
                # skip the 'Base: 2011-12' benchmark-year banner, not a data column
                if "base" in v.lower():
                    continue
                mobj = fy_re.search(v)
                if mobj:
                    label = f"{mobj.group(1)}-{mobj.group(2)}"
                    cols.setdefault(label, c)
        if len(cols) > len(best_cols):
            best_cols = cols
    return best_cols


def read_state_series(path, sheet):
    """Read one 'State/Union Territory' sheet -> {fy_label: {raw_name: float}} and
    an ordered list of the raw state names encountered.

    Returns (by_fy, raw_names). Non-region / aggregate rows are left in raw_names so the
    caller can report skips; value coercion drops '-'.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    hdr, name_col = find_header(ws)
    ycols = year_columns(ws, hdr, name_col)
    by_fy = {fy: {} for fy in ycols}
    raw_names = []
    for r in range(hdr + 1, ws.max_row + 1):
        nm = ws.cell(row=r, column=name_col).value
        if not isinstance(nm, str):
            continue
        nm = nm.strip()
        if not nm:
            continue
        # stop at footnote rows (they start with '-:' , '*:' , 'Note', 'Source', digits map)
        low = nm.lower()
        if low.startswith(("-:", "*:", "note", "source", "sources", "sdls")):
            continue
        raw_names.append(nm)
        for fy, c in ycols.items():
            val = num(ws.cell(row=r, column=c).value)
            if val is not None:
                by_fy[fy][nm] = val
    return by_fy, raw_names, sorted(ycols.keys())


def read_t176_total(path):
    """Latest end-March Outstanding-Liabilities total per state from T176.

    T176 has 2 sheets per year (components part 1 + '(Concld.)' part 2). The total lives
    on the '(Concld.)' sheet in the last column labelled 'Outstanding Liabilities'
    (formula '20 = (14 to 19)'). We pick the sheet whose title carries the largest year.
    Returns (year:int, {raw_name: float}, raw_names, sheet_title).
    """
    import re
    wb = openpyxl.load_workbook(path, data_only=True)
    best = None  # (year, sheet_title)
    for sn in wb.sheetnames:
        ws = wb[sn]
        title = ws.cell(row=2, column=2).value
        if not isinstance(title, str) or "concld" not in title.lower():
            continue
        mobj = re.search(r"-\s*(\d{4})", title)
        if not mobj:
            continue
        yr = int(mobj.group(1))
        if best is None or yr > best[0]:
            best = (yr, sn)
    if best is None:
        raise RuntimeError("no T176 '(Concld.)' sheet with a year found")
    yr, sn = best
    ws = wb[sn]
    hdr, name_col = find_header(ws)
    # locate the 'Outstanding Liabilities' total column on the header row
    total_col = None
    for c in range(name_col + 1, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if isinstance(v, str) and "outstanding" in v.lower() and "liabilit" in v.lower():
            total_col = c
    if total_col is None:
        # fallback: the very last data column
        total_col = ws.max_column
    vals, raw_names = {}, []
    for r in range(hdr + 1, ws.max_row + 1):
        nm = ws.cell(row=r, column=name_col).value
        if not isinstance(nm, str):
            continue
        nm = nm.strip()
        low = nm.lower()
        if not nm or low.startswith(("sdls", "note", "source")):
            continue
        raw_names.append(nm)
        val = num(ws.cell(row=r, column=total_col).value)
        if val is not None:
            vals[nm] = val
    return yr, vals, raw_names, sn


def resolve(m, raw_name):
    """RBI raw name -> region_keys state code (via STATE_ALIASES then RegionMatcher).
    Returns None for aggregates and unmatched names."""
    if raw_name.strip().lower() in AGG:
        return None
    canon = STATE_ALIASES.get(raw_name.strip().lower(), raw_name)
    return m.state_code(canon)


def latest_gsdp_by_state(gsdp_by_fy, gsdp_years):
    """For each state code, {code: (fy_label, gsdp_lakh)} at its own latest usable FY,
    plus prev-year for growth. gsdp_by_fy is keyed by fy then raw_name."""
    # invert to raw_name -> {fy: val}
    per_name = {}
    for fy, d in gsdp_by_fy.items():
        for nm, v in d.items():
            per_name.setdefault(nm, {})[fy] = v
    return per_name  # caller resolves names + picks year


def main():
    con = sqlite3.connect(DB)
    m = RegionMatcher(con)

    metric_ids = (
        "econ_percapita_nsdp_rbi",
        "gsdp_growth",
        "fiscal_deficit_pct_gsdp",
        "own_tax_pct_gsdp",
        "outstanding_debt_pct_gsdp",
    )
    qs = ",".join("?" * len(metric_ids))
    con.execute(f"DELETE FROM metric_values WHERE metric_id IN ({qs})", metric_ids)

    # ---- Load raw tables -------------------------------------------------
    t19_i, _, _ = read_state_series(os.path.join(RAW, FILES["T19"]), "T_19(i)")
    t19_ii, t19_names, t19_years_ii = read_state_series(os.path.join(RAW, FILES["T19"]), "T_19(ii)")
    t21_i, _, _ = read_state_series(os.path.join(RAW, FILES["T21"]), "T_21(i)")
    t21_ii, t21_names, t21_years_ii = read_state_series(os.path.join(RAW, FILES["T21"]), "T_21(ii)")
    t164_ii, t164_names, t164_years = read_state_series(os.path.join(RAW, FILES["T164"]), "T_164(ii)")
    t168_ii, t168_names, t168_years = read_state_series(os.path.join(RAW, FILES["T168"]), "T_168(ii)")
    t176_year, t176_vals, t176_names, t176_sheet = read_t176_total(os.path.join(RAW, FILES["T176"]))

    # merge (i)+(ii) per-name maps for T19 and T21 (full FY history)
    def merge_series(part_i, part_ii):
        per_name = {}
        for part in (part_i, part_ii):
            for fy, d in part.items():
                for nm, v in d.items():
                    per_name.setdefault(nm, {})[fy] = v
        return per_name

    t19_name_fy = merge_series(t19_i, t19_ii)
    t21_name_fy = merge_series(t21_i, t21_ii)

    def sorted_fys(name_fy):
        allfy = set()
        for d in name_fy.values():
            allfy |= set(d.keys())
        return sorted(allfy)  # ascending fiscal years

    t21_all_fys = sorted_fys(t21_name_fy)  # e.g. ['2011-12',...,'2024-25']

    skipped = {k: [] for k in ("percapita", "growth", "deficit", "owntax", "debt")}

    # ---- 1) econ_percapita_nsdp_rbi : latest usable FY per state (T19) ----
    percapita = {}          # code -> (fy_label, value)
    for nm, fymap in t19_name_fy.items():
        code = resolve(m, nm)
        if code is None:
            if nm.strip().lower() not in AGG:
                skipped["percapita"].append(f"{nm} (no region code)")
            continue
        usable = [fy for fy in t21_all_fys if fymap.get(fy) is not None]
        if not usable:
            skipped["percapita"].append(f"{nm} (no numeric value)")
            continue
        fy = usable[-1]
        percapita[code] = (fy, fymap[fy])

    # ---- 2) gsdp_growth : two latest FYs with values (T21) ----
    growth = {}             # code -> (fy_label_latest, pct)
    for nm, fymap in t21_name_fy.items():
        code = resolve(m, nm)
        if code is None:
            if nm.strip().lower() not in AGG:
                skipped["growth"].append(f"{nm} (no region code)")
            continue
        usable = [fy for fy in t21_all_fys if fymap.get(fy) is not None]
        if len(usable) < 2:
            skipped["growth"].append(f"{nm} (need 2 GSDP years, have {len(usable)})")
            continue
        latest, prev = usable[-1], usable[-2]
        g = (fymap[latest] / fymap[prev] - 1.0) * 100.0
        growth[code] = (latest, g)

    # helper: GSDP (in lakh) for a state at a given FY, by resolving names
    # build code -> {fy: gsdp_lakh}
    gsdp_by_code = {}
    for nm, fymap in t21_name_fy.items():
        code = resolve(m, nm)
        if code is None:
            continue
        gsdp_by_code.setdefault(code, {}).update(fymap)

    def ratio_metric(name_fy_num, numer_years_desc, skipkey):
        """Generic: numerator table (code->{fy:val_crore}) / GSDP(crore) x100 at the
        latest FY where BOTH exist. numer_years_desc: FY labels newest-first.
        Returns {code: (fy_label, pct)}."""
        # build code -> {fy: numer}
        num_by_code = {}
        for nm, fymap in name_fy_num.items():
            code = resolve(m, nm)
            if code is None:
                if nm.strip().lower() not in AGG:
                    skipped[skipkey].append(f"{nm} (no region code)")
                continue
            num_by_code.setdefault(code, {}).update(fymap)
        out = {}
        for code, fymap in num_by_code.items():
            g = gsdp_by_code.get(code, {})
            # candidate FYs: those present in numerator, newest first
            cand = [fy for fy in numer_years_desc if fymap.get(fy) is not None]
            chosen = None
            for fy in cand:
                if g.get(fy) is not None and g[fy] != 0:
                    chosen = fy
                    break
            if chosen is None:
                skipped[skipkey].append(f"code {code} (no common FY with GSDP)")
                continue
            gsdp_crore = g[chosen] / 100.0
            pct = fymap[chosen] / gsdp_crore * 100.0
            out[code] = (chosen, pct)
        return out

    # numerator FY history for T164/T168 (from (ii) sheet only; that's where recent FYs live)
    def to_name_fy(part):
        pn = {}
        for fy, d in part.items():
            for nm, v in d.items():
                pn.setdefault(nm, {})[fy] = v
        return pn

    t164_name_fy = to_name_fy(t164_ii)
    t168_name_fy = to_name_fy(t168_ii)

    fy_desc = list(reversed(t21_all_fys))  # newest-first list of all FYs

    # ---- 3) fiscal_deficit_pct_gsdp ----
    deficit = ratio_metric(t164_name_fy, fy_desc, "deficit")

    # ---- 4) own_tax_pct_gsdp ----
    owntax = ratio_metric(t168_name_fy, fy_desc, "owntax")

    # ---- 5) outstanding_debt_pct_gsdp ----
    # T176 stock is at end-March <year> == close of FY (year-1)-(year). end-March 2025 => FY2024-25.
    debt_fy = f"{t176_year-1}-{str(t176_year)[2:]}"      # 2025 -> '2024-25'
    debt = {}
    for nm, v in t176_vals.items():
        code = resolve(m, nm)
        if code is None:
            if nm.strip().lower() not in AGG:
                skipped["debt"].append(f"{nm} (no region code)")
            continue
        g = gsdp_by_code.get(code, {})
        # match the debt stock year to GSDP of the same FY; fall back to prior FY if missing
        chosen = None
        for fy in [debt_fy] + [f for f in fy_desc if f != debt_fy]:
            if g.get(fy):
                chosen = fy
                break
        if chosen is None:
            skipped["debt"].append(f"{nm} (no GSDP for debt ratio)")
            continue
        gsdp_crore = g[chosen] / 100.0
        debt[code] = (chosen, v / gsdp_crore * 100.0)

    # ---- upsert metric rows -------------------------------------------------
    meth_units = ("Unit alignment: T21 GSDP is in Rs Lakh; T164/T168/T176 are in Rs Crore, "
                  "so GSDP_crore = T21_lakh / 100 before forming the ratio. Ratios are computed "
                  "at the latest fiscal year for which both the numerator table and T21 GSDP "
                  "carry a value for that state.")

    upsert_metric(
        con, "econ_percapita_nsdp_rbi",
        "Per-capita NSDP (RBI HBS, latest FY)", "economy", "₹", 0, 1,
        "Per-capita Net State Domestic Product at current prices, latest available fiscal "
        "year, from RBI Handbook of Statistics on Indian States 2025 (Table 19). Distinct "
        "from the MoSPI `econ_percapita_nsdp` series; this is the RBI FY-latest compilation "
        "(mostly 2024-25, a few states 2023-24 where 2024-25 is not yet available).",
        SOURCE, URL, LICENSE, 2024,
        methodology=("RBI HBS 2025 Table 19 (Per Capita NSDP, current prices, base 2011-12), "
                     "per state its latest published fiscal year (2024-25 where available, "
                     "else 2023-24). Underlying figures sourced by RBI from NSO/MoSPI. "
                     "State/UT level only."))

    upsert_metric(
        con, "gsdp_growth",
        "GSDP growth (nominal, YoY)", "economy", "%", 1, 1,
        "Year-on-year nominal growth of Gross State Domestic Product at current prices, "
        "from RBI Handbook of Statistics on Indian States 2025 (Table 21), computed over "
        "the two latest fiscal years available for each state (mostly 2023-24 -> 2024-25).",
        SOURCE, URL, LICENSE, 2024,
        methodology=("gsdp_growth = (GSDP_latestFY / GSDP_prevFY - 1) x 100, from RBI HBS 2025 "
                     "Table 21 (GSDP current prices, Rs Lakh). Nominal (current-price) growth. "
                     "Two latest FYs with a value per state; for most states 2024-25 over "
                     "2023-24, for states without a 2024-25 figure 2023-24 over 2022-23."))

    upsert_metric(
        con, "fiscal_deficit_pct_gsdp",
        "Gross fiscal deficit (% of GSDP)", "economy", "%", 1, 0,
        "State gross fiscal deficit as a share of GSDP, from RBI Handbook of Statistics on "
        "Indian States 2025: Gross Fiscal Deficit (Table 164) divided by GSDP (Table 21). "
        "Latest fiscal year is 2024-25 (Budget Estimates) for most states.",
        SOURCE, URL, LICENSE, 2024,
        methodology=("fiscal_deficit_pct_gsdp = GFD (T164, Rs Crore) / (GSDP T21 Rs Lakh / 100) x 100. "
                     "Latest common FY per state: 2024-25 is Budget Estimates (BE), 2023-24 Revised "
                     "Estimates (RE), 2022-23 Actuals (A). Most states use 2024-25 (BE). A negative "
                     "value means a fiscal surplus. " + meth_units))

    upsert_metric(
        con, "own_tax_pct_gsdp",
        "Own tax revenue (% of GSDP)", "economy", "%", 1, 1,
        "State own tax revenue as a share of GSDP, from RBI Handbook of Statistics on Indian "
        "States 2025: Own Tax Revenue (Table 168) divided by GSDP (Table 21). Latest fiscal "
        "year is 2024-25 (Budget Estimates) for most states.",
        SOURCE, URL, LICENSE, 2024,
        methodology=("own_tax_pct_gsdp = Own Tax Revenue (T168, Rs Crore) / (GSDP T21 Rs Lakh / 100) x 100. "
                     "Latest common FY per state: 2024-25 (BE), 2023-24 (RE), 2022-23 (A). Most states "
                     "use 2024-25 (BE). Measures a state's tax effort / fiscal self-reliance. " + meth_units))

    upsert_metric(
        con, "outstanding_debt_pct_gsdp",
        "Outstanding liabilities (% of GSDP)", "economy", "%", 1, 0,
        "State outstanding liabilities (total debt stock) as a share of GSDP, from RBI Handbook "
        "of Statistics on Indian States 2025: Outstanding Liabilities total (Table 176, as at "
        f"end-March {t176_year}) divided by GSDP (Table 21). The debt stock at end-March {t176_year} "
        f"corresponds to the close of fiscal year {debt_fy}.",
        SOURCE, URL, LICENSE, 2024,
        methodology=("outstanding_debt_pct_gsdp = Outstanding Liabilities total (T176 'Outstanding "
                     "Liabilities' column, formula 20 = (14 to 19), Rs Crore, as at end-March "
                     f"{t176_year}) / (GSDP T21 Rs Lakh / 100) x 100. The end-March {t176_year} stock is "
                     f"matched to GSDP of FY {debt_fy}; where that GSDP is unavailable the nearest prior "
                     "FY GSDP is used. Debt figures from RBI 'State Finances: A Study of Budgets'. "
                     + meth_units))

    # ---- write state values (estimated flag set for BE/RE years) ------------
    def est_flag(fy_label):
        tag = ESTIMATE_TAG.get(fy_label)
        return 1 if tag in ("BE", "RE") else 0

    def write(mid, code_to_fyval, use_est=True):
        n = 0
        for code, (fy, val) in code_to_fyval.items():
            yr = fy_end_year(fy)
            est = est_flag(fy) if use_est else 0
            # 'projected', never 'inherited': a BE/RE figure is the state's own
            # projection for a fiscal year that has not closed. Nothing is copied
            # from another region and there is no donor to cite (adr-021).
            con.execute("INSERT OR REPLACE INTO metric_values"
                        "(metric_id,region_code,region_level,year,value,estimated,estimate_kind) "
                        "VALUES(?,?,?,?,?,?,?)",
                        (mid, code, "state", yr, round(float(val), 4), est,
                         "projected" if est else None))
            n += 1
        return n

    n_pc = write("econ_percapita_nsdp_rbi", percapita, use_est=False)   # NSO figures, not BE/RE tagged
    n_gr = write("gsdp_growth", growth, use_est=False)
    n_fd = write("fiscal_deficit_pct_gsdp", deficit, use_est=True)
    n_ot = write("own_tax_pct_gsdp", owntax, use_est=True)
    # debt stock (end-March) is provisional but not tagged BE/RE; flag 0
    n_db = write("outstanding_debt_pct_gsdp", debt, use_est=False)

    total = n_pc + n_gr + n_fd + n_ot + n_db
    notes = (f"RBI HBS 2025 state fiscal vertical (5 metrics, economy, state-level). "
             f"percapita_nsdp_rbi={n_pc} (T19 latest FY per state); "
             f"gsdp_growth={n_gr} (T21 two latest FYs); "
             f"fiscal_deficit_pct_gsdp={n_fd} (T164/T21, mostly 2024-25 BE); "
             f"own_tax_pct_gsdp={n_ot} (T168/T21, mostly 2024-25 BE); "
             f"outstanding_debt_pct_gsdp={n_db} (T176 end-March {t176_year} / T21). "
             f"GSDP unit-converted lakh/100 -> crore. Skips: {skipped}.")
    log_load(con, "ingest_rbi_fiscal.py", SOURCE, 2024, LICENSE, FETCHED, total, notes)

    con.commit()
    con.close()

    print(f"WROTE {total} state values:")
    print(f"  econ_percapita_nsdp_rbi  = {n_pc}")
    print(f"  gsdp_growth              = {n_gr}")
    print(f"  fiscal_deficit_pct_gsdp  = {n_fd}")
    print(f"  own_tax_pct_gsdp         = {n_ot}")
    print(f"  outstanding_debt_pct_gsdp= {n_db}")
    print(f"T176 total column used: sheet {t176_sheet!r}, end-March {t176_year} (FY {debt_fy})")
    for k, v in skipped.items():
        if v:
            print(f"SKIPPED[{k}]: {v}")


if __name__ == "__main__":
    main()
